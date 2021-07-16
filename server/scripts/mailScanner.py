import sys
from openpyxl import load_workbook
# Import to access driver to interact with browser
from selenium import webdriver
# Import to access common keys to give instructions to browser
from selenium.webdriver.common.keys import Keys
# Imports to implement waiting for elements to be present before executing the rest of the script
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Path the the chrome webdriver which we downloaded from google and saved locally
PATH = "C:\Program Files (x86)\chromedriver.exe"

# Store driver as variable, specify browser, in this case Chrome
driver = webdriver.Chrome(PATH)

print("Starting MailScanner ... ")

excel_path = sys.argv[1].split(',')[0]

print(f"Paths passed: {excel_path}")

def run_automation(path):
    driver.get("https://google.com")

    wb = load_workbook(filename=path, read_only=True)
    ws = wb.active
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data_item = {}
        for headers in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            i = 0
            for header in headers:
                data_item[header] = row[i]
                i+=1
        data.append(data_item)

    print(data)

    for item in data:
        search = driver.find_element_by_css_selector('input[name="q"]')
        for key, val in item.items():
            search.send_keys(val)

run_automation(excel_path)